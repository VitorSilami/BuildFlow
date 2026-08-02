import io
from decimal import Decimal

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from buildflow.configuracoes.eap_import import MAX_EAP_IMPORT_BYTES
from buildflow.configuracoes.eap_import import ArquivoInvalido
from buildflow.configuracoes.eap_import import LinhasInvalidas
from buildflow.configuracoes.eap_import import importar_eap_de_arquivo
from buildflow.configuracoes.models import CatalogoServico
from buildflow.configuracoes.models import Disciplina
from buildflow.configuracoes.models import Unidade
from buildflow.registros_diarios.tests.factories import DisciplinaFactory
from buildflow.registros_diarios.tests.factories import ProjetoParaRdoFactory
from buildflow.registros_diarios.tests.factories import UnidadeFactory

pytestmark = pytest.mark.django_db


def _csv_upload(nome: str, conteudo: str, encoding: str = "utf-8") -> SimpleUploadedFile:
    return SimpleUploadedFile(nome, conteudo.encode(encoding), content_type="text/csv")


def _xlsx_upload(nome: str, linhas: list[list]) -> SimpleUploadedFile:
    workbook = Workbook()
    sheet = workbook.active
    for linha in linhas:
        sheet.append(linha)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return SimpleUploadedFile(
        nome,
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def test_csv_com_cabecalho_na_primeira_linha_importa_corretamente():
    projeto = ProjetoParaRdoFactory()
    csv_texto = (
        "DISCIPLINA,ATIVIDADE,UN,TOTAL\n"
        "Terraplenagem,Corte,m3,1500\n"
        "Terraplenagem,Aterro,m3,800\n"
    )
    arquivo = _csv_upload("import.csv", csv_texto)

    resultado = importar_eap_de_arquivo(projeto, arquivo)

    assert resultado.disciplinas_criadas == 1
    assert resultado.servicos_criados == 2
    disciplina = Disciplina.objects.get(projeto=projeto, nome="Terraplenagem")
    assert disciplina.pai is None
    assert set(disciplina.servicos.values_list("nome", flat=True)) == {"Corte", "Aterro"}
    corte = disciplina.servicos.get(nome="Corte")
    assert corte.quantidade_planejada == Decimal("1500")
    assert corte.peso_percentual is None
    assert corte.unidade.sigla == "m3"


def test_csv_com_linhas_de_titulo_antes_do_cabecalho_e_importado():
    projeto = ProjetoParaRdoFactory()
    csv_texto = (
        "MODELO IMPORT SOFT\n"
        "Lote 2 - Patrocinio\n"
        "\n"
        "CHAVE,EAP,DISCIPLINA,ATIVIDADE,UN,TOTAL\n"
        "1,1,Terraplenagem,Corte,m3,1500\n"
        "2,2,Terraplenagem,Aterro,m3,800\n"
    )
    arquivo = _csv_upload("legado.csv", csv_texto)

    resultado = importar_eap_de_arquivo(projeto, arquivo)

    assert resultado.disciplinas_criadas == 1
    assert resultado.servicos_criados == 2


def test_xlsx_com_mesmo_conteudo_produz_mesmo_resultado():
    projeto = ProjetoParaRdoFactory()
    arquivo = _xlsx_upload(
        "import.xlsx",
        [
            ["DISCIPLINA", "ATIVIDADE", "UN", "TOTAL"],
            ["Terraplenagem", "Corte", "m3", 1500],
            ["Terraplenagem", "Aterro", "m3", 800],
        ],
    )

    resultado = importar_eap_de_arquivo(projeto, arquivo)

    assert resultado.disciplinas_criadas == 1
    assert resultado.servicos_criados == 2


def test_csv_em_cp1252_com_acentuacao_decodifica_corretamente():
    projeto = ProjetoParaRdoFactory()
    csv_texto = (
        "DISCIPLINA,ATIVIDADE,UN,TOTAL\n"
        "Terraplenagem,Preparação de subleito,m3,500\n"
    )
    arquivo = _csv_upload("import.csv", csv_texto, encoding="cp1252")

    resultado = importar_eap_de_arquivo(projeto, arquivo)

    assert resultado.servicos_criados == 1
    servico = CatalogoServico.objects.get(disciplina__projeto=projeto)
    assert servico.nome == "Preparação de subleito"


def test_linha_com_disciplina_em_branco_gera_erro_e_nao_cria_nada():
    projeto = ProjetoParaRdoFactory()
    csv_texto = (
        "DISCIPLINA,ATIVIDADE,UN,TOTAL\n"
        "Terraplenagem,Corte,m3,1500\n"
        ",Aterro,m3,800\n"
    )
    arquivo = _csv_upload("import.csv", csv_texto)

    with pytest.raises(LinhasInvalidas) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert exc_info.value.erros == ["Linha 3: DISCIPLINA em branco."]
    assert not Disciplina.objects.filter(projeto=projeto).exists()


def test_linha_com_total_nao_numerico_gera_erro_e_nao_cria_nada():
    projeto = ProjetoParaRdoFactory()
    csv_texto = "DISCIPLINA,ATIVIDADE,UN,TOTAL\nTerraplenagem,Corte,m3,abc\n"
    arquivo = _csv_upload("import.csv", csv_texto)

    with pytest.raises(LinhasInvalidas) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert exc_info.value.erros == ["Linha 2: TOTAL/QUANTIDADE inválido."]
    assert not Disciplina.objects.filter(projeto=projeto).exists()


def test_atividade_duplicada_na_mesma_disciplina_gera_erro():
    projeto = ProjetoParaRdoFactory()
    csv_texto = (
        "DISCIPLINA,ATIVIDADE,UN,TOTAL\n"
        "Terraplenagem,Corte,m3,1500\n"
        "Terraplenagem,Corte,m3,200\n"
    )
    arquivo = _csv_upload("import.csv", csv_texto)

    with pytest.raises(LinhasInvalidas) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert exc_info.value.erros == ["Linha 3: ATIVIDADE duplicada para esta DISCIPLINA."]


def test_multiplas_linhas_com_mesma_disciplina_agrupam_em_uma_disciplina():
    projeto = ProjetoParaRdoFactory()
    csv_texto = (
        "DISCIPLINA,ATIVIDADE,UN,TOTAL\n"
        "Terraplenagem,Corte,m3,1500\n"
        "Terraplenagem,Aterro,m3,800\n"
        "Drenagem,Escavação de vala,m,200\n"
    )
    arquivo = _csv_upload("import.csv", csv_texto)

    resultado = importar_eap_de_arquivo(projeto, arquivo)

    assert resultado.disciplinas_criadas == 2
    assert resultado.servicos_criados == 3
    terraplenagem = Disciplina.objects.get(projeto=projeto, nome="Terraplenagem")
    assert terraplenagem.servicos.count() == 2


def test_projeto_com_eap_nao_vazia_rejeita_import():
    projeto = ProjetoParaRdoFactory()
    DisciplinaFactory(projeto=projeto)
    arquivo = _csv_upload("import.csv", "DISCIPLINA,ATIVIDADE,UN,TOTAL\nTerraplenagem,Corte,m3,1500\n")

    with pytest.raises(ArquivoInvalido) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert "já possui uma EAP" in exc_info.value.mensagem


def test_extensao_nao_suportada_e_rejeitada():
    projeto = ProjetoParaRdoFactory()
    arquivo = SimpleUploadedFile("import.txt", b"conteudo", content_type="text/plain")

    with pytest.raises(ArquivoInvalido) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert "não suportado" in exc_info.value.mensagem


def test_cabecalho_nao_encontrado_e_rejeitado():
    projeto = ProjetoParaRdoFactory()
    arquivo = _csv_upload("import.csv", "col1,col2\nvalor1,valor2\n")

    with pytest.raises(ArquivoInvalido) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert "Cabeçalho não encontrado" in exc_info.value.mensagem


def test_arquivo_maior_que_limite_e_rejeitado():
    projeto = ProjetoParaRdoFactory()
    conteudo = b"x" * (MAX_EAP_IMPORT_BYTES + 1)
    arquivo = SimpleUploadedFile("import.csv", conteudo, content_type="text/csv")

    with pytest.raises(ArquivoInvalido) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert "tamanho máximo" in exc_info.value.mensagem


def test_xlsx_com_cabecalho_em_aba_nao_ativa_e_importado():
    projeto = ProjetoParaRdoFactory()
    workbook = Workbook()
    aba_vazia = workbook.active
    aba_vazia.title = "EXPORT_PROJECT"
    aba_vazia.append(["algo", "irrelevante"])

    aba_dados = workbook.create_sheet("BASE_QTD_L2")
    aba_dados.append(["DISCIPLINA", "ATIVIDADE", "UN", "TOTAL"])
    aba_dados.append(["Terraplenagem", "Corte", "m3", 1500])

    workbook.active = 0

    buffer = io.BytesIO()
    workbook.save(buffer)
    arquivo = SimpleUploadedFile(
        "import.xlsx",
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    resultado = importar_eap_de_arquivo(projeto, arquivo)

    assert resultado.disciplinas_criadas == 1
    assert resultado.servicos_criados == 1


def test_unidade_com_sigla_diferente_por_caixa_reaproveita_unidade_existente():
    projeto = ProjetoParaRdoFactory()
    existente = UnidadeFactory(sigla="m3")
    csv_texto = "DISCIPLINA,ATIVIDADE,UN,TOTAL\nTerraplenagem,Corte,M3,1500\n"
    arquivo = _csv_upload("import.csv", csv_texto)

    importar_eap_de_arquivo(projeto, arquivo)

    assert Unidade.objects.filter(sigla__iexact="m3").count() == 1
    servico = CatalogoServico.objects.get(disciplina__projeto=projeto)
    assert servico.unidade_id == existente.id


def test_unidade_muito_longa_gera_erro():
    projeto = ProjetoParaRdoFactory()
    unidade_longa = "x" * 17
    csv_texto = f"DISCIPLINA,ATIVIDADE,UN,TOTAL\nTerraplenagem,Corte,{unidade_longa},1500\n"
    arquivo = _csv_upload("import.csv", csv_texto)

    with pytest.raises(LinhasInvalidas) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert exc_info.value.erros == [
        "Linha 2: UNIDADE excede o tamanho máximo (16 caracteres).",
    ]


def test_quantidade_negativa_gera_erro():
    projeto = ProjetoParaRdoFactory()
    csv_texto = "DISCIPLINA,ATIVIDADE,UN,TOTAL\nTerraplenagem,Corte,m3,-100\n"
    arquivo = _csv_upload("import.csv", csv_texto)

    with pytest.raises(LinhasInvalidas) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert exc_info.value.erros == ["Linha 2: TOTAL/QUANTIDADE inválido."]


def test_quantidade_acima_do_limite_gera_erro():
    projeto = ProjetoParaRdoFactory()
    csv_texto = "DISCIPLINA,ATIVIDADE,UN,TOTAL\nTerraplenagem,Corte,m3,9999999999\n"
    arquivo = _csv_upload("import.csv", csv_texto)

    with pytest.raises(LinhasInvalidas) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert exc_info.value.erros == ["Linha 2: TOTAL/QUANTIDADE inválido."]


def test_disciplina_muito_longa_gera_erro():
    projeto = ProjetoParaRdoFactory()
    nome_longo = "D" * 256
    csv_texto = f"DISCIPLINA,ATIVIDADE,UN,TOTAL\n{nome_longo},Corte,m3,1500\n"
    arquivo = _csv_upload("import.csv", csv_texto)

    with pytest.raises(LinhasInvalidas) as exc_info:
        importar_eap_de_arquivo(projeto, arquivo)

    assert exc_info.value.erros == [
        "Linha 2: DISCIPLINA excede o tamanho máximo (255 caracteres).",
    ]


def test_linha_com_nota_em_coluna_ignorada_e_tratada_como_branco():
    projeto = ProjetoParaRdoFactory()
    csv_texto = (
        "CHAVE,EAP,DISCIPLINA,ATIVIDADE,UN,TOTAL\n"
        "1,1,Terraplenagem,Corte,m3,1500\n"
        "Nota: quantidades parametrizadas, substituir quando consolidadas.,,,,,\n"
    )
    arquivo = _csv_upload("import.csv", csv_texto)

    resultado = importar_eap_de_arquivo(projeto, arquivo)

    assert resultado.disciplinas_criadas == 1
    assert resultado.servicos_criados == 1


def test_arquivo_com_formato_hierarquico_e_detectado_automaticamente():
    projeto = ProjetoParaRdoFactory()
    arquivo = _xlsx_upload(
        "import.xlsx",
        [
            ["CÓDIGO", "Task Name", "UNIDADE", "QUANTIDADE"],
            ["001", "Mobilização", "", ""],
            ["001.001", "Canteiro", "vb", 1],
        ],
    )

    resultado = importar_eap_de_arquivo(projeto, arquivo)

    assert resultado.disciplinas_criadas == 1
    assert resultado.servicos_criados == 1
