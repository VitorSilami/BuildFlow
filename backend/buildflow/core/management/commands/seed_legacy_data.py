from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.core.management.base import CommandError
from django.db import transaction

from buildflow.configuracoes.models import CatalogoServico
from buildflow.configuracoes.models import Disciplina
from buildflow.configuracoes.models import Unidade
from buildflow.empresas.models import Empresa
from buildflow.projetos.models import Projeto
from buildflow.usuarios.models import PerfilChoices
from buildflow.usuarios.models import User

PLANILHA_PADRAO = "MODELO IMPORT SOFT (1).xlsx"
ABA_QUANTIDADES = "BASE_QTD_L2"
LINHA_CABECALHO = 4  # CHAVE | EAP | DISCIPLINA | ATIVIDADE | UN | TOTAL


class Command(BaseCommand):
    help = (
        "Importa BASE_QTD_L2 da planilha MODELO IMPORT SOFT como Disciplina/"
        "CatalogoServico (com peso/quantidade) de um projeto de demonstracao "
        "(dado legado usado so para carga inicial — ver decisao registrada em "
        "memoria de projeto, nao define schema)."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--planilha",
            default=None,
            help="Caminho da planilha .xlsx",
        )

    def handle(self, *args, **options):
        caminho = Path(
            options["planilha"] or settings.BASE_DIR.parent / PLANILHA_PADRAO,
        )
        if not caminho.exists():
            msg = f"Planilha nao encontrada em {caminho}"
            raise CommandError(msg)

        workbook = openpyxl.load_workbook(caminho, data_only=True)
        if ABA_QUANTIDADES not in workbook.sheetnames:
            msg = f"Aba '{ABA_QUANTIDADES}' nao encontrada na planilha."
            raise CommandError(msg)

        with transaction.atomic():
            projeto = self._get_or_create_projeto_legado()
            linhas = self._ler_linhas(workbook[ABA_QUANTIDADES])
            self._importar_disciplinas_e_servicos(projeto, linhas)

        self.stdout.write(
            self.style.SUCCESS(
                f"Dados legados importados para o projeto '{projeto.nome}'.",
            ),
        )

    def _get_or_create_projeto_legado(self) -> Projeto:
        empresa, _ = Empresa.objects.get_or_create(
            slug="obra-lote-2-patrocinio",
            defaults={"nome": "Obra Lote 2 — Patrocínio (dados legados)"},
        )
        gerente, criado = User.objects.get_or_create(
            email="gerente@obra-lote-2-patrocinio.buildflow.local",
            defaults={
                "nome": "Gerente Lote 2",
                "empresa": empresa,
                "perfil": PerfilChoices.GERENTE,
                "is_active": True,
            },
        )
        if criado:
            gerente.set_unusable_password()
            gerente.save(update_fields=["password"])

        projeto, _ = Projeto.objects.get_or_create(
            empresa=empresa,
            nome="Lote 2 — Patrocínio (EAP importada)",
            defaults={
                "descricao": (
                    "Projeto de demonstração populado a partir de "
                    "MODELO IMPORT SOFT (BASE_QTD_L2)."
                ),
                "criado_por": gerente,
            },
        )
        return projeto

    def _ler_linhas(self, planilha) -> list[dict]:
        linhas = []
        for row in planilha.iter_rows(min_row=LINHA_CABECALHO + 1, values_only=True):
            chave, _eap, disciplina, atividade, unidade_sigla, total = row[:6]
            if not disciplina or not atividade:
                continue
            linhas.append(
                {
                    "chave": chave,
                    "disciplina": str(disciplina).strip(),
                    "atividade": str(atividade).strip(),
                    "unidade": str(unidade_sigla).strip() if unidade_sigla else "un",
                    "total": float(total) if total is not None else 0.0,
                },
            )
        return linhas

    def _importar_disciplinas_e_servicos(
        self,
        projeto: Projeto,
        linhas: list[dict],
    ) -> None:
        for linha in linhas:
            unidade, _ = Unidade.objects.get_or_create(sigla=linha["unidade"])
            disciplina, _ = Disciplina.objects.get_or_create(
                projeto=projeto,
                nome=linha["disciplina"],
            )
            CatalogoServico.objects.update_or_create(
                disciplina=disciplina,
                nome=linha["atividade"],
                defaults={
                    "unidade": unidade,
                    "quantidade_planejada": linha["total"],
                },
            )
